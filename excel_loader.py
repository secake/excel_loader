import json
import re
import copy
from openpyxl import load_workbook
from django.contrib.postgres.fields import ArrayField

from log import Log


class SheetInfo:

    def __init__(self, info: dict):
        self.name = info.get('sheet_name')
        # 标题行
        self.header_line = int(info.get('header_line', 1))
        # 从哪行开始解析
        self.start_line = int(info.get('start_line', 1))
        # 需要忽略的值
        self.ignore_values = info.get('ignore_values', [])


class FieldMap:
    
    # 在存储时，当唯一键发生冲突时的解决方案
    conflict_map = {'ignore': 0, 'replace': 1}
    def __init__(self, map: dict):
        self.headers = map.get('headers')
        self.values = map.get('values', {})
        self.final_values = map.get('final', {})
        self.func = map.get('func', [])
        self.required = map.get('required', False)
        self.unique = map.get('unique', False)
        self.conflict = FieldMap.conflict_map[map.get('conflict', 'ignore')]


class ClassMap:
    def __init__(self, maps: dict, header_map):
        self.cls_map = {}
    
        # eg. k: Student.age
        # eg. map: { "headers": [xx,xx], "values": [xx,xx]}
        for k, map in maps.items():
            cls_str = k.split('.', maxsplit=1)[0]
            field_str = k.split('.', maxsplit=1)[1]
            if cls_str not in self.cls_map:
                self.cls_map[cls_str] = {}
            self.cls_map[cls_str][field_str] = FieldMap(map)
            for header in self.cls_map[cls_str][field_str].headers:
                if header not in header_map:
                    header_map[header] = [{'cls': cls_str, 'field': field_str}]
                else:
                    header_map[header].append({'cls': cls_str, 'field': field_str})

    @classmethod
    def init(cls, maps, header_map):
        return ClassMap(maps, header_map).cls_map

class Loader:
    
    def config(self, field, expected_type=dict):
        """
        获取配置  
        field: 要获取的字段  
        expected_type: 期望的类型，默认dict  
        """

        # 如果第一次调用，加载config。 热加载
        
        # 传入的config 是json格式字符串
        if not isinstance(self.__config, dict):
            try:
                self.__config = json.loads(self.__config)
            except:
                pass
        
        # 传入的config 是存储json文本的文件路径
        if not isinstance(self.__config, dict):
            try:
                self.__config = json.load(open(self.__config,'r'))
            except:
                raise Exception(self.log.err('config 加载失败'))

        # 获取某个 config列
        if field in self.__config and isinstance(self.__config[field], expected_type):
            return self.__config[field]
        else:
            raise Exception(self.log.err(
                'json config 字段{}存在问题， 类型必须为{}'.format(field, expected_type)))


    def __init__(self, config, path:str='', log=Log(), globals=globals()):
        """
        config: json形式的字符串， 或存储json文本的文件路径  
        path: excel文件路径
        log: 指定日志的接收者   
        globals: 指定当前全局类环境  
        """

        self.__config = config
        self.log = log
        self.globals = globals

        # 反序列化配置
        self.__header_map = {}
        self.__sheets = [SheetInfo(sheet) for sheet in self.config('sheets', list)]
        self.__cls_map = ClassMap.init(self.config('maps'), self.__header_map)

        # 校验类是否存在
        for cls_str in self.__cls_map.keys():
            if cls_str not in self.globals:
                raise Exception(self.log.err('当前命名空间不存在类：{}'.format(cls_str)))
            # 校验列是否存在
            obj = self.globals[cls_str]()
            for field_str in self.__cls_map[cls_str].keys():
                if not hasattr(obj, field_str):
                    raise Exception(self.log.err('类:{}不存在列:{}'.format(cls_str, field_str)))

        # 从文件中加载的数据
        self.__workbook = None
        self.__objs = []

        self.log.info('config 文件加载成功！')
        if path:
            self.load(path)


    # 可多次调用
    def load(self, path):
        self.log.info('开始加载文件{}'.format(path))
        wb = load_workbook(path)
        
        # for sheet in self.__sheets:
        #     if sheet.name not in wb:
        #         raise Exception(self.log.err('{}中不存在表格{}'.format(path, sheet.name)))
        self.__workbook = wb
        for sheet in self.__sheets:
            self.__load_sheet(sheet)

    def __load_sheet(self, sheet: SheetInfo):
        """
        加载一个表格
        """
        self.log.info('- 开始加载表格{}'.format(sheet.name))
        ws = self.__workbook[sheet.name]
        headers = [h.value for h in ws[sheet.header_line]]
        for row in list(ws.rows)[sheet.start_line:]:
            self.__load_row(headers, row, sheet.ignore_values)

    
    def __load_row(self, headers, row, ignore_values):
        """
        加载一行数据
        """
        num = 0
        data_without_func = {}
        data = {cls_str: {} for cls_str in self.__cls_map.keys()}
        for cell in row:
            value = cell.value
            num += 1
            # 被忽略的值，如 none
            if value in ignore_values:
                continue
            if len(headers) < num:
                break

            # 获取header
            header = headers[num-1]
            if header not in self.__header_map:
                continue
            
            # 先一行值全部加载进来
            data_without_func[header] = value

        def call_func(self, data, header, value):
            """
            调用指定的函数函数
            """
            if not value:
                return
            for cls_field in self.__header_map[header]:
                cls_str = cls_field['cls']
                field_str = cls_field['field']

                func = self.__cls_map[cls_str][field_str].func

                if not func:
                    data[cls_str][field_str] = value

                if func == 'listappend':
                    if field_str not in data[cls_str] or not isinstance(data[cls_str][field_str], list):
                        data[cls_str][field_str] = []
                    data[cls_str][field_str].append(value)
                if func == 'strappend':
                    if field_str not in data[cls_str] or not isinstance(data[cls_str][field_str], str):
                        data[cls_str][field_str] = ''
                    data[cls_str][field_str] += str(value)
                if func == 'setadd':
                    if field_str not in data[cls_str] or not isinstance(data[cls_str][field_str], set):
                        data[cls_str][field_str] = set()
                    data[cls_str][field_str].add(value)
                if func == 'numadd':
                    if field_str not in data[cls_str] \
                            or type(data[cls_str][field_str]) not in [int, float]:
                        data[cls_str][field_str] = 0.0
                    data[cls_str][field_str] += float(value)


        for h, v in data_without_func.items():
            call_func(self, data, h, v)
        self.log.debug('-- 加载了一行数据: {}'.format(json.dumps(data)))

        dest_data = {}
        # 进行值的映射，解析，校验等
        for cls_str, field_maps in self.__cls_map.items():
            dest_data[cls_str] = {}
            for field_str, field_map in field_maps.items():
                self.__load_value(data, dest_data, cls_str, field_str)
        self.log.debug('-- 校验了一行数据: {}'.format(json.dumps(dest_data)))

        # 如果这一行值有问题，跳过这一行，处理下一行
        objs = {cls_str:self.globals[cls_str]() for cls_str in dest_data.keys()}
        if not self.__verity_row(dest_data, objs):
            self.log.warn('-- 跳过了当前行， 必填此段解析失败, ', [obj.__dict__ for obj in objs.values()])
            return
        
        self.log.info('-- 当前行成功加载', [obj.__dict__ for obj in objs.values()])
        # 加载过的所有数据  
        self.__objs.append(objs)

    def __load_value(self, src_data, dest_data, cls_str, field_str):
        '''
         $f:() extra_field，是否从其他字段进行映射，没有这个参数不进行映射
         $r:() 验证是否符合正则表达式， 没有这个参数不进行正则校验  
         $g:() 需要同时存在$r:()，意思取值正则的哪一个分组， 没有这个参数取group()分组 
         $d:   default, 缺省值，当没有在给定的值中出现，赋值为缺省值，没有这个参数的话，缺失值是下面的 None, 如果field_map的values是空，缺失值是下面的$s
         $s:   self, 即将值赋值成本身
        '''
        value = src_data[cls_str].get(field_str, None)
        def get_params(value):
            """
            是否包含特殊处理参数
            """
            src_cls = cls_str
            src_field = field_str
            re_pattern = re.compile(r'.*')
            re_group = 0
    
            has_param = False
            # 是否指定了从外部列映射
            re_search = re.search(r'\$f:\('
                                  r'(?P<src_cls>[\w]*?[^\\])\.'
                                  r'(?P<src_field>[\w]*?[^\\])\)',
                                  value)
            if re_search \
                    and re_search.group('src_cls') in src_data \
                    and re_search.group('src_field') in src_data[re_search.group('src_cls')]:
                src_cls = re_search.group('src_cls')
                src_field = re_search.group('src_field')
                has_param = True

            # 可以指定取值的哪一部分
            re_search = re.search(r'\$r:\('
                                  r'(?P<pattern>.*?[^\\]|)\)',
                                  value)
            if re_search:
                re_pattern = re.compile(re_search.group('pattern'))
                has_param = True

            # 可以指定取正则匹配到的哪一个分组，默认 group(0)
            re_search = re.search(r'\$g:\('
                                  r'(?P<group>[\d]*?)\)',
                                  value)
            if re_search:
                re_group = int(re_search.group('group'))
            return src_cls, src_field, re_pattern, re_group, has_param


        def is_default_param(value):
            # 是否指定了缺省值
            re_search =  re.search(r'\$d', value)
            if re_search:
                return True
            return False
                
        field_map = self.__cls_map[cls_str][field_str]
        # 缺省值
        default_value = '$s' if not field_map.values else None
        
        # #
        # 找到值的映射
        # #
        for src_value, dest_value in field_map.values.items():
            # 如果值匹配，直接结束
            if value == src_value:
                default_value = dest_value
                break
            
            # 如果匹配了特殊参数，直接返结束
            src_cls, src_field, re_pattern, re_group, has_param  = get_params(src_value)
            if has_param:
                re_search = re_pattern.search(str(src_data[src_cls][src_field]))
                if not re_search or not re_search.group(re_group):
                    continue
                else:
                    default_value = dest_value
                    break
            
            # 如果是$s，直接结束
            if src_value == '$s':
                default_value = dest_value
                break

            # 如果找到了缺省值，设置缺省值
            if is_default_param(src_value):
                default_value = dest_value
                continue


        # #
        # 进行值的映射
        # #

        # 设置匹配自身
        if default_value == '$s':
            dest_data[cls_str][field_str] = value
            return dest_data
        
        # 如果没有找到匹配，并且field_map非空
        if default_value == None:
            dest_data[cls_str][field_str] = None
            return dest_data

        # 如果匹配了特殊参数，进行映射
        src_cls, src_field, re_pattern, re_group, has_param  = get_params(default_value)
        if has_param:
            re_search = re_pattern.search(str(src_data[src_cls][src_field]))
            if re_search:
                dest_data[cls_str][field_str] = re_search.group(re_group)
                return dest_data
        
        # field_map非空，并且找到了匹配
        dest_data[cls_str][field_str] = default_value
        return dest_data


    def __verity_row(self, data, objs):
        """
        进行一行的数据校验 和 最终的值处理
        """
        
        for cls_str, field_maps in self.__cls_map.items():
            for field_str, field_map in field_maps.items():
                # 类型校验
                field_type = type(getattr(objs[cls_str], field_str))
                value = data[cls_str].get(field_str, None)
                if value != None:
                    try:
                        if field_type in [list, dict, ArrayField]:
                            value = json.loads(value)
                        value = field_type(value)
                    except:
                        value = None
                # 必填性校验
                if value == None:
                    if field_map.required == True:
                        return False
                    continue
                value = field_type(value)
                setattr(objs[cls_str], field_str, value)
        return True


    def out_json_str(self):
        """
        返回加载后的数据 的json 字符串形式
        """
        ret = []
        for objs in self.__objs:
            ret.append({k: obj.__dict__ for k, obj in objs.items()})
        return json.dumps(ret)

    def out_objs(self):
        return self.__objs

    def save(self):
        """
        存储到库中，谨慎调用
        """
        unique_fileds = {cls_str:{'fields':[], 'conflict': 'ignore'} for cls_str in self.__cls_map.keys()}
        # 是否存在唯一键
        for cls_str, field_maps in self.__cls_map.items():
            for field_str, field_map in field_maps:
                if field_map.unique == True:
                    unique_fileds[cls_str]['fields'].append(field_str)
                if field_map.conflict == 'replace':
                    unique_fileds[cls_str]['conflict'] = 'replace'

        for objs in self.__objs:
            for cls_str, obj in objs.items():
                alread_exist = False
                # 查找是否有唯一键冲突的数据
                for unique_filed in unique_fileds[cls_str]['fields']:
                    db_obj = self.globals[cls_str].objects.filter(unique_field=getattr(obj, unique_filed)).first()
                    if db_obj:
                        alread_exist = True
                        # 更新所有字段
                        if unique_fileds[cls_str]['conflict'] == 'replace':
                            self.log.warn('库中已存在', cls_str, '{}={}'.format(unique_field, getattr(obj, unique_fileds)), '的数据，根据配置规则，已数据替换到库中')
                            obj.id = db_obj.id
                            obj.save()
                        # 忽略修改，不保存
                        else:
                            self.log.info('库中已存在', cls_str, '{}={}'.format(unique_field, getattr(obj, unique_fileds)), '的数据，根据配置规则，已将此次修改忽略')
                            break
                # 新增
                if not alread_exist:
                    self.log.info('库中不存在', cls_str, '{}={}'.format(unique_field, getattr(obj, unique_fileds)), '的数据，已新增')        
                    obj.save()
